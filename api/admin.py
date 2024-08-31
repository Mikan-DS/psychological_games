import csv
from datetime import timedelta

import openpyxl
from django.contrib import admin
from django.forms import inlineformset_factory
from django.http import HttpResponse
from django.urls import reverse
from django.utils.html import format_html
from django.utils.timezone import make_naive
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from django.utils.translation import gettext_lazy as _

from .models import ProjectSummary, ProjectParameter, TestResult, TestResultParameter


class ProjectParameterInline(admin.TabularInline):
    model = ProjectParameter
    extra = 1


class TestResultInline(admin.TabularInline):
    model = TestResult
    extra = 1
    fields = ['name', 'ip', 'end_time', 'duration', 'view_test_result']
    readonly_fields = ['view_test_result']

    def view_test_result(self, instance):
        if instance.id:
            url = reverse('admin:api_testresult_change', args=[instance.id])
            return format_html('<a href="{}">Просмотр</a>', url)
        return ""

    view_test_result.short_description = "Результаты"


class ProjectSummaryAdmin(admin.ModelAdmin):
    inlines = [ProjectParameterInline, TestResultInline]
    list_display = ['name']
    search_fields = ['name']


admin.site.register(ProjectSummary, ProjectSummaryAdmin)


def export_to_csv(modeladmin, request, queryset):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="test_results.csv"'
    response.write('\ufeff'.encode('utf-8'))  # BOM (Byte Order Mark)

    writer = csv.writer(response)
    writer.writerow(['Проект', 'Имя', 'IP', 'Дата завершения', 'Время в игре'] +
                    [param.name for param in ProjectParameter.objects.all()])

    for obj in queryset:
        row = [obj.project, obj.name, obj.ip, obj.end_time, obj.duration]
        parameters = {param.name: param.value for param in obj.parameters.all()}
        for param in ProjectParameter.objects.all():
            row.append(parameters.get(param.name, ''))
        writer.writerow(row)

    return response


export_to_csv.short_description = "Экспорт в CSV (разрабатывается)"


def export_to_excel(modeladmin, request, queryset):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="test_results.xlsx"'

    wb = openpyxl.Workbook()

    if queryset.count() >= 1:

        wb.remove(wb.active)

        for project in ProjectSummary.objects.all():
            create_excel_sheet(wb, project, queryset)

    wb.save(response)
    return response


def create_excel_sheet(workbook, project: ProjectSummary, queryset):
    project_results = queryset.filter(project=project)

    if project_results.count() < 1:
        return

    ws = workbook.create_sheet()
    ws.title = project.name

    ws.freeze_panes = 'A2'

    parameters = project.parameters.all()

    headers = ['Тариф', 'Номер телефона', 'Имя', 'IP', 'Дата завершения', 'Время в игре'] + [param.name for param in
                                                                                             parameters]
    ws.append(headers)
    for i, _ in enumerate(headers):
        ws.cell(row=1, column=i + 1).font = Font(bold=True)

    for obj in project.results.all():
        end_time = obj.end_time.strftime("%Y-%m-%d %H:%M:%S")

        tarif = 'Анонимный'
        phone = ''
        if obj.user:
            tarif = "Зарегистрирован"
            if obj.user.username.isnumeric():
                phone = "+" + obj.user.username

            purchases = obj.user.purchases.filter(paid=True)
            if purchases:
                tarif = purchases.last().item_type

        name = obj.name
        if obj.user:
            name = " ".join((obj.user.first_name, obj.user.last_name)) or obj.user.username

        row = [tarif, phone, name, obj.ip, end_time, str(timedelta(seconds=obj.duration))]
        for param in parameters:
            obj_param = obj.parameters.filter(project_parameter=param)
            if obj_param:
                obj_param = obj_param.first()
                if obj_param and obj_param.value.isnumeric():
                    row.append(int(obj_param.value))
                else:
                    row.append(obj_param.value)
            else:
                row.append("")
        ws.append(row)

    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width


export_to_excel.short_description = "Экспорт в Excel"


class TestResultParameterInline(admin.TabularInline):
    model = TestResultParameter
    extra = 1
    fields = ['name', 'value']


class PaidFilter(admin.SimpleListFilter):
    title = _('Статус оплаты')
    parameter_name = 'paid_status'

    def lookups(self, request, model_admin):
        return (
            ('paid', _('Платные')),
            ('free', _('Бесплатные')),
        )

    def queryset(self, request, queryset):
        if self.value() == 'paid':
            return queryset.filter(user__purchases__paid=True).distinct()
        if self.value() == 'free':
            return queryset.exclude(user__purchases__paid=True)

class TestResultAdmin(admin.ModelAdmin):
    inlines = [TestResultParameterInline]
    list_display = ['project', 'user', 'name', 'ip', 'end_time', 'duration']
    search_fields = ['project__name', 'name', 'ip', 'user__username']
    list_filter = ['project', 'end_time', PaidFilter]

    actions = [export_to_excel]


admin.site.register(TestResult, TestResultAdmin)


admin.site.register(ProjectParameter, admin.ModelAdmin)
admin.site.register(TestResultParameter, admin.ModelAdmin)
